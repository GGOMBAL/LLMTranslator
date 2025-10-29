#!/usr/bin/env python3
"""
번역 결과를 읽기 좋은 Excel 형태로 변환하는 스크립트
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_translation_data(json_path):
    """JSON 파일에서 번역 데이터 로드"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def create_bilingual_excel(data, output_path):
    """이중 언어(원문-번역) Excel 파일 생성"""

    # 페이지별 데이터를 DataFrame으로 변환
    pages_data = []
    for page in data['pages']:
        pages_data.append({
            '페이지': page['page_number'],
            '원문 (中文)': page['original_text'],
            '번역문 (English)': page['translated_text'],
            '상태': '성공' if '[Translation failed' not in page['translated_text'] and '[TOC -' not in page['translated_text'] else '실패',
            '원문 글자수': page['original_char_count'],
            '번역문 글자수': page['translated_char_count'],
            '처리 시간(초)': round(page['translation_time'], 2)
        })

    df_pages = pd.DataFrame(pages_data)

    # 통계 데이터 생성
    stats_data = {
        '항목': [
            '총 처리 페이지',
            '성공적인 번역',
            '실패한 번역',
            '성공률 (%)',
            '처리 시간',
            '평균 페이지당 처리 시간(초)'
        ],
        '값': [
            data['total_pages_processed'],
            data['successful_translations'],
            data['total_pages_processed'] - data['successful_translations'],
            f"{(data['successful_translations'] / data['total_pages_processed'] * 100):.1f}%",
            data['timestamp'],
            f"{df_pages['처리 시간(초)'].mean():.2f}"
        ]
    }
    df_stats = pd.DataFrame(stats_data)

    # 성공한 페이지만 필터링
    df_success = df_pages[df_pages['상태'] == '성공'].copy()

    # 실패한 페이지만 필터링
    df_failed = df_pages[df_pages['상태'] == '실패'].copy()

    # Excel 파일 생성
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. 통계 시트
        df_stats.to_excel(writer, sheet_name='📊 번역 통계', index=False)

        # 2. 전체 페이지 시트
        df_pages.to_excel(writer, sheet_name='📄 전체 페이지', index=False)

        # 3. 성공한 번역 시트
        if not df_success.empty:
            df_success.to_excel(writer, sheet_name='✅ 성공 번역', index=False)

        # 4. 실패한 번역 시트
        if not df_failed.empty:
            df_failed.to_excel(writer, sheet_name='❌ 실패 목록', index=False)

    # 스타일 적용
    apply_excel_formatting(output_path)

    return output_path

def apply_excel_formatting(excel_path):
    """Excel 파일에 포맷팅 적용"""

    wb = load_workbook(excel_path)

    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 각 시트에 대해 포맷팅 적용
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 헤더 스타일
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)

        # 첫 번째 행(헤더)에 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        # 텍스트 길이 계산 (한글은 2배 가중)
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)
                except:
                    pass

            # 최대/최소 너비 설정
            adjusted_width = min(max(max_length, 10), 80)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 데이터 셀 스타일
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # 상태 열에 색상 적용
                if ws.cell(1, cell.column).value == '상태':
                    if cell.value == '성공':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == '실패':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)

        # 행 높이 자동 조정
        ws.row_dimensions[1].height = 30  # 헤더 높이

        # 틀 고정 (헤더 행)
        ws.freeze_panes = 'A2'

    # 저장
    wb.save(excel_path)

def main():
    """메인 실행 함수"""

    # 입력/출력 파일 경로
    json_path = '/Users/kimjinmyung/Desktop/GIT/TranslatewithRag/final_translation_results.json'
    output_path = '/Users/kimjinmyung/Desktop/GIT/TranslatewithRag/번역결과_읽기좋은버전.xlsx'

    print("🔄 번역 데이터를 읽는 중...")
    data = load_translation_data(json_path)

    print("📝 Excel 파일 생성 중...")
    result_path = create_bilingual_excel(data, output_path)

    print(f"\n✅ 완료!")
    print(f"📁 저장 위치: {result_path}")
    print(f"\n📊 통계:")
    print(f"  - 총 페이지: {data['total_pages_processed']}")
    print(f"  - 성공: {data['successful_translations']}")
    print(f"  - 실패: {data['total_pages_processed'] - data['successful_translations']}")
    print(f"  - 성공률: {(data['successful_translations'] / data['total_pages_processed'] * 100):.1f}%")

if __name__ == '__main__':
    main()
