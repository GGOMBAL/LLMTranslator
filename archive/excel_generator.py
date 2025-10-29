import os
import logging
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class ExcelGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_simple_excel(self, pages_text: List[Tuple[int, str]], output_path: str, 
                           title: str = "Translated Document") -> str:
        """Create a simple Excel file from translated pages"""
        try:
            # Create DataFrame
            data = []
            for page_num, text in pages_text:
                if text.strip():
                    # Split text into paragraphs
                    paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
                    for i, paragraph in enumerate(paragraphs):
                        data.append({
                            'Page': page_num,
                            'Paragraph': i + 1,
                            'Text': paragraph
                        })
            
            df = pd.DataFrame(data)
            
            # Save to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Translation', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Translation']
                self._format_simple_worksheet(worksheet, title)
            
            self.logger.info(f"Simple Excel document saved to: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error creating simple Excel document: {str(e)}")
            raise
    
    def create_bilingual_excel(self, original_chunks: List[Dict], translated_chunks: List[Dict], 
                              output_path: str, title: str = "Bilingual Translation") -> str:
        """Create a bilingual Excel file with original and translated text"""
        try:
            data = []
            
            for orig_chunk, trans_chunk in zip(original_chunks, translated_chunks):
                for (orig_page_num, orig_text), (trans_page_num, trans_text) in zip(
                    orig_chunk['pages'], trans_chunk['pages']
                ):
                    if orig_text.strip() or trans_text.strip():
                        # Split texts into paragraphs
                        orig_paragraphs = [p.strip() for p in orig_text.split('\n') if p.strip()]
                        trans_paragraphs = [p.strip() for p in trans_text.split('\n') if p.strip()]
                        
                        # Match paragraphs (pad shorter list with empty strings)
                        max_paragraphs = max(len(orig_paragraphs), len(trans_paragraphs))
                        orig_paragraphs.extend([''] * (max_paragraphs - len(orig_paragraphs)))
                        trans_paragraphs.extend([''] * (max_paragraphs - len(trans_paragraphs)))
                        
                        for i, (orig_para, trans_para) in enumerate(zip(orig_paragraphs, trans_paragraphs)):
                            data.append({
                                'Page': orig_page_num,
                                'Paragraph': i + 1,
                                'Original (Chinese)': orig_para,
                                'Translation (English)': trans_para
                            })
            
            df = pd.DataFrame(data)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Bilingual Translation', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Bilingual Translation']
                self._format_bilingual_worksheet(worksheet, title)
            
            self.logger.info(f"Bilingual Excel document saved to: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error creating bilingual Excel document: {str(e)}")
            raise
    
    def create_structured_excel(self, translated_chunks: List[Dict], output_path: str,
                               title: str = "Structured Translation") -> str:
        """Create a structured Excel file with multiple sheets"""
        try:
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, translated_chunks, title)
            
            # Create translation sheet
            translation_sheet = workbook.create_sheet("Translation")
            self._create_translation_sheet(translation_sheet, translated_chunks)
            
            # Create statistics sheet
            stats_sheet = workbook.create_sheet("Statistics")
            self._create_statistics_sheet(stats_sheet, translated_chunks)
            
            # Save workbook
            workbook.save(output_path)
            
            self.logger.info(f"Structured Excel document saved to: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error creating structured Excel document: {str(e)}")
            raise
    
    def _format_simple_worksheet(self, worksheet: Worksheet, title: str):
        """Format simple worksheet"""
        # Set title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Merge title row
        worksheet.merge_cells('A1:C1')
        
        # Add headers in row 3
        headers = ['Page', 'Paragraph', 'Text']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 80
        
        # Set text wrapping for text column
        for row in worksheet.iter_rows(min_row=4):
            if row[2].value:
                row[2].alignment = Alignment(wrap_text=True, vertical='top')
    
    def _format_bilingual_worksheet(self, worksheet: Worksheet, title: str):
        """Format bilingual worksheet"""
        # Set title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Merge title row
        worksheet.merge_cells('A1:D1')
        
        # Add headers in row 3
        headers = ['Page', 'Paragraph', 'Original (Chinese)', 'Translation (English)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 40
        worksheet.column_dimensions['D'].width = 40
        
        # Set text wrapping for text columns
        for row in worksheet.iter_rows(min_row=4):
            if len(row) >= 4:
                for cell in row[2:4]:  # Original and Translation columns
                    if cell.value:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_summary_sheet(self, worksheet: Worksheet, chunks: List[Dict], title: str):
        """Create summary sheet"""
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=18, bold=True)
        
        total_pages = sum(len(chunk['pages']) for chunk in chunks)
        total_chars = sum(sum(len(text) for _, text in chunk['pages']) for chunk in chunks)
        
        summary_data = [
            ['Total Chunks', len(chunks)],
            ['Total Pages', total_pages],
            ['Total Characters', total_chars],
            ['Average Characters per Page', total_chars // total_pages if total_pages > 0 else 0]
        ]
        
        for row, (label, value) in enumerate(summary_data, 3):
            worksheet[f'A{row}'] = label
            worksheet[f'B{row}'] = value
            worksheet[f'A{row}'].font = Font(bold=True)
    
    def _create_translation_sheet(self, worksheet: Worksheet, chunks: List[Dict]):
        """Create main translation sheet"""
        worksheet['A1'] = 'Translation Content'
        worksheet['A1'].font = Font(size=16, bold=True)
        
        headers = ['Page', 'Text']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        row = 4
        for chunk in chunks:
            for page_num, text in chunk['pages']:
                if text.strip():
                    worksheet[f'A{row}'] = page_num
                    worksheet[f'B{row}'] = text
                    worksheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 80
    
    def _create_statistics_sheet(self, worksheet: Worksheet, chunks: List[Dict]):
        """Create statistics sheet"""
        worksheet['A1'] = 'Translation Statistics'
        worksheet['A1'].font = Font(size=16, bold=True)
        
        headers = ['Chunk', 'Start Page', 'End Page', 'Pages Count', 'Total Characters', 'Avg Chars/Page']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        for i, chunk in enumerate(chunks, 1):
            row = 3 + i
            total_chars = sum(len(text) for _, text in chunk['pages'])
            pages_count = len(chunk['pages'])
            avg_chars = total_chars // pages_count if pages_count > 0 else 0
            
            worksheet[f'A{row}'] = i
            worksheet[f'B{row}'] = chunk['start_page']
            worksheet[f'C{row}'] = chunk['end_page']
            worksheet[f'D{row}'] = pages_count
            worksheet[f'E{row}'] = total_chars
            worksheet[f'F{row}'] = avg_chars
        
        # Auto-adjust column widths
        for col in range(1, 7):
            worksheet.column_dimensions[chr(64 + col)].width = 15
    
    def generate_excel_filename(self, input_pdf_path: str, suffix: str = "_translated") -> str:
        """Generate Excel filename based on input PDF path"""
        base_name = os.path.splitext(os.path.basename(input_pdf_path))[0]
        output_dir = os.path.dirname(input_pdf_path)
        return os.path.join(output_dir, f"{base_name}{suffix}.xlsx")