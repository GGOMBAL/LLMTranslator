#!/usr/bin/env python3
"""
읽기 좋은 Excel 번역 결과 생성기

기능:
- 줄바꿈 자동 인식 및 표시
- 한자(간체) 완벽 표시 (UTF-8)
- 자동 열 너비 조정
- 색상 코딩 (성공/실패)
- 헤더 스타일링
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def load_json_results(json_path: str) -> dict:
    """JSON 결과 파일 로드"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ 파일을 찾을 수 없습니다: {json_path}")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ JSON 파싱 에러: {e}")
        return None

def create_readable_excel(data: dict, output_path: str):
    """읽기 좋은 Excel 파일 생성"""

    # 워크북 생성
    wb = openpyxl.Workbook()

    # 시트 1: 통계
    ws_stats = wb.active
    ws_stats.title = "📊 통계"
    create_statistics_sheet(ws_stats, data)

    # 시트 2: 전체 번역 결과
    ws_all = wb.create_sheet("📄 전체 페이지")
    create_translation_sheet(ws_all, data['pages'])

    # 시트 3: 성공만
    successful = [p for p in data['pages'] if not p['translated_text'].startswith('[')]
    ws_success = wb.create_sheet("✅ 성공")
    create_translation_sheet(ws_success, successful)

    # 시트 4: 실패/처리 필요
    failed = [p for p in data['pages'] if p['translated_text'].startswith('[')]
    if failed:
        ws_failed = wb.create_sheet("⚠️ 실패")
        create_translation_sheet(ws_failed, failed)

    # 저장
    try:
        wb.save(output_path)
        print(f"✅ Excel 파일 생성 완료: {output_path}")
        return True
    except Exception as e:
        print(f"❌ Excel 저장 실패: {e}")
        return False

def create_statistics_sheet(ws, data):
    """통계 시트 생성"""

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)

    # 제목
    ws['A1'] = "📊 번역 결과 통계"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')

    # 통계 데이터
    stats = [
        ("", ""),
        ("총 페이지 수", data['total_pages_processed']),
        ("성공한 번역", data['successful_translations']),
        ("실패/처리", data['total_pages_processed'] - data['successful_translations']),
        ("성공률", f"{data['successful_translations']/data['total_pages_processed']*100:.1f}%"),
        ("처리 시간", data.get('timestamp', 'N/A')),
        ("버전", data.get('version', 'V2')),
    ]

    row = 2
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value

        # 라벨 스타일
        if label:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        row += 1

    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

def create_translation_sheet(ws, pages):
    """번역 결과 시트 생성"""

    # 헤더
    headers = ['페이지', '원문 샘플', '번역문', '원문 길이', '번역문 길이', '상태', '시간(초)']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 데이터 작성
    for row_idx, page in enumerate(pages, 2):
        page_num = page['page_number']
        original = page['original_text']
        translated = page['translated_text']

        # 상태 판단
        is_success = not translated.startswith('[')
        status = "✅ 성공" if is_success else "⚠️ 실패"

        # 셀 값 설정
        ws.cell(row=row_idx, column=1, value=page_num)
        ws.cell(row=row_idx, column=2, value=original[:300])  # 원문 샘플
        ws.cell(row=row_idx, column=3, value=translated[:300] if len(translated) > 300 else translated)
        ws.cell(row=row_idx, column=4, value=page['original_char_count'])
        ws.cell(row=row_idx, column=5, value=page['translated_char_count'])
        ws.cell(row=row_idx, column=6, value=status)
        ws.cell(row=row_idx, column=7, value=page.get('translation_time', 'N/A'))

        # 스타일 적용
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border

            # 텍스트 래핑 (줄바꿈 표시)
            if col in [2, 3]:  # 원문, 번역문 컬럼
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 성공/실패에 따른 색상
        status_cell = ws.cell(row=row_idx, column=6)
        if is_success:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")

    # 열 너비 자동 조정
    ws.column_dimensions['A'].width = 8   # 페이지
    ws.column_dimensions['B'].width = 50  # 원문
    ws.column_dimensions['C'].width = 50  # 번역문
    ws.column_dimensions['D'].width = 12  # 원문 길이
    ws.column_dimensions['E'].width = 12  # 번역문 길이
    ws.column_dimensions['F'].width = 12  # 상태
    ws.column_dimensions['G'].width = 10  # 시간

    # 행 높이 설정 (줄바꿈 고려)
    for row in range(2, len(pages) + 2):
        ws.row_dimensions[row].height = 80  # 충분한 높이

def main():
    """메인 함수"""
    print("=" * 80)
    print("📊 읽기 좋은 Excel 번역 결과 생성기")
    print("=" * 80)

    # V2 결과 처리
    v2_json = "output/improved_translation_v2_results.json"
    if os.path.exists(v2_json):
        print(f"\n📄 V2 결과 처리 중...")
        data = load_json_results(v2_json)
        if data:
            output_excel = "output/improved_translation_v2_results_readable.xlsx"
            create_readable_excel(data, output_excel)
            print(f"✅ V2 Excel: {output_excel}")
    else:
        print(f"\n⚠️  V2 결과 파일이 없습니다: {v2_json}")

    # V3 결과 처리
    v3_json = "output/improved_translation_v3_results.json"
    if os.path.exists(v3_json):
        print(f"\n📄 V3 결과 처리 중...")
        data = load_json_results(v3_json)
        if data:
            output_excel = "output/improved_translation_v3_results_readable.xlsx"
            create_readable_excel(data, output_excel)
            print(f"✅ V3 Excel: {output_excel}")
    else:
        print(f"\n⚠️  V3 결과 파일이 없습니다: {v3_json}")

    print("\n✅ 완료!")
    print("=" * 80)

if __name__ == "__main__":
    main()
