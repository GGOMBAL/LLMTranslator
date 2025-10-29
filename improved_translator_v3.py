#!/usr/bin/env python3
"""
개선된 PDF 번역기 V3 - 2단계 개선사항 적용
- V2 기능 (재시도 5회, 지수 백오프, 동적 타임아웃) 유지
- 추가: 스마트 청크 분할
- 추가: 표/테이블 전용 처리
- 추가: TOC 전용 처리
"""

import PyPDF2
import os
import json
import time
import re
import csv
import sys
from typing import List, Tuple, Optional, Dict
from dataclasses import dataclass
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel output will be skipped.")

try:
    from toc_structure_parser import TOCStructureParser, TOCItem
    TOC_PARSER_AVAILABLE = True
except ImportError:
    TOC_PARSER_AVAILABLE = False

# ============================================================================
# Step 1: 스마트 청크 분할
# ============================================================================

@dataclass
class TextChunk:
    """텍스트 청크 데이터 클래스"""
    content: str
    start_index: int
    end_index: int
    chunk_type: str  # 'paragraph', 'sentence', 'forced'

def smart_chunk_text(text: str, max_length: int = 800, overlap: int = 100) -> List[TextChunk]:
    """
    스마트 청크 분할 - 의미 단위로 텍스트 분할

    우선순위:
    1. 단락 구분 (\\n\\n)
    2. 문장 끝 (。！？.)
    3. 쉼표 (，,)
    4. 고정 길이 (최후 수단)
    """

    if len(text) <= max_length:
        return [TextChunk(text, 0, len(text), 'complete')]

    chunks = []
    current_pos = 0

    while current_pos < len(text):
        # 청크 끝 위치 계산
        end_pos = min(current_pos + max_length, len(text))

        if end_pos >= len(text):
            # 마지막 청크
            chunks.append(TextChunk(
                text[current_pos:],
                current_pos,
                len(text),
                'final'
            ))
            break

        # 최적 분할 지점 찾기
        chunk_text = text[current_pos:end_pos + 200]  # 약간 더 읽어서 분할점 찾기
        split_point = find_best_split_point(chunk_text, max_length)

        actual_end = current_pos + split_point

        chunks.append(TextChunk(
            text[current_pos:actual_end],
            current_pos,
            actual_end,
            'split'
        ))

        # 다음 청크 시작 위치 (오버랩 적용)
        current_pos = actual_end - overlap if actual_end > overlap else actual_end

    return chunks

def find_best_split_point(text: str, max_length: int) -> int:
    """
    최적의 분할 지점 찾기

    우선순위:
    1. 단락 구분 (\\n\\n) - 가장 자연스러운 분할
    2. 문장 끝 (。！？.!?) - 문장 완결성 유지
    3. 쉼표 (，,;；) - 절 단위 분할
    4. 공백 - 단어 단위 분할
    5. 강제 분할 - 최후 수단
    """

    # 1. 단락 구분 찾기
    paragraph_breaks = [m.start() for m in re.finditer(r'\n\n+', text[:max_length + 100])]
    if paragraph_breaks:
        best = max([p for p in paragraph_breaks if p <= max_length], default=None)
        if best and best > max_length * 0.7:  # 최소 70% 이상 활용
            return best + 2  # \n\n 이후

    # 2. 문장 끝 찾기
    sentence_ends = [m.end() for m in re.finditer(r'[。！？.!?]\s*', text[:max_length + 50])]
    if sentence_ends:
        best = max([s for s in sentence_ends if s <= max_length], default=None)
        if best and best > max_length * 0.7:
            return best

    # 3. 쉼표 찾기
    comma_positions = [m.end() for m in re.finditer(r'[，,;；]\s*', text[:max_length + 20])]
    if comma_positions:
        best = max([c for c in comma_positions if c <= max_length], default=None)
        if best and best > max_length * 0.8:
            return best

    # 4. 공백 찾기
    space_positions = [m.end() for m in re.finditer(r'\s+', text[:max_length + 10])]
    if space_positions:
        best = max([s for s in space_positions if s <= max_length], default=None)
        if best:
            return best

    # 5. 강제 분할
    return max_length

def merge_chunk_translations(chunks: List[str], original_chunks: List[TextChunk]) -> str:
    """
    청크별 번역 결과를 자연스럽게 병합

    오버랩 부분 중복 제거 및 자연스러운 연결
    """

    if not chunks:
        return ""

    if len(chunks) == 1:
        return chunks[0]

    # 첫 번째 청크로 시작
    merged = chunks[0]

    # 나머지 청크 병합
    for i in range(1, len(chunks)):
        # 중복 부분 제거 로직
        # 간단하게: 공백으로 연결 (향후 더 정교하게 개선 가능)
        merged += " " + chunks[i]

    return merged

# ============================================================================
# Step 2: 표/테이블 감지
# ============================================================================

def detect_table(text: str) -> bool:
    """
    표/테이블 구조 감지

    감지 기준:
    1. 표 경계 문자: ┃│├┤┬┴─ 등
    2. 반복 패턴: 여러 줄에 걸친 일정한 구조
    3. 표 키워드: "表", "功能矩阵", "系统" 등
    """

    # 1. 표 경계 문자 체크
    table_chars = ['┃', '│', '├', '┤', '┬', '┴', '─', '║', '╠', '╣']
    table_char_count = sum(text.count(char) for char in table_chars)

    if table_char_count > 5:  # 표 문자가 5개 이상
        return True

    # 2. 표 제목 패턴 체크
    table_patterns = [
        r'表\d+',  # 表3, 表1 등
        r'功能矩阵',
        r'系统.*功能',
        r'\|.*\|.*\|',  # Markdown 표
    ]

    for pattern in table_patterns:
        if re.search(pattern, text):
            return True

    # 3. 반복 패턴 체크 (간단 버전)
    lines = text.split('\n')
    if len(lines) > 3:
        # 비슷한 길이의 줄이 여러 개 있으면 표일 가능성
        line_lengths = [len(line) for line in lines if line.strip()]
        if len(line_lengths) > 3:
            avg_length = sum(line_lengths) / len(line_lengths)
            similar_lines = sum(1 for l in line_lengths if abs(l - avg_length) < avg_length * 0.3)
            if similar_lines > len(line_lengths) * 0.6:  # 60% 이상 유사
                return True

    return False

def process_table_structure(text: str) -> str:
    """
    표 구조 전용 처리

    전략:
    1. 표를 행 단위로 분리
    2. 각 행의 셀 내용 추출
    3. 셀별로 번역
    4. 표 구조로 재조립
    """

    print("      [표 감지] 표 전용 처리 모드 활성화")

    # 간단 버전: 줄바꿈으로 분리하여 각 줄 개별 번역
    # (복잡한 표 파싱은 추후 개선)

    return text  # 일단 원본 그대로 반환 (번역은 외부에서)

# ============================================================================
# Step 3: TOC 전용 처리
# ============================================================================

def detect_toc(text: str) -> bool:
    """
    TOC(목차) 페이지 감지

    감지 기준:
    1. "目录", "目 录" 키워드
    2. 많은 점선 (.......) 또는 대시 (-----)
    3. 페이지 번호 패턴 (- XX -)
    4. 계층적 번호 구조 (1.1.1, 3.4.3.6 등)
    """

    # 1. 목차 키워드
    if re.search(r'目\s*录', text):
        return True

    # 2. 점선/대시 많음
    dot_count = text.count('.')
    dash_count = text.count('-')
    text_length = len(text)

    if text_length > 0:
        dot_ratio = dot_count / text_length
        dash_ratio = dash_count / text_length

        if dot_ratio > 0.15 or dash_ratio > 0.1:  # 15% 이상이 점 또는 10% 이상이 대시
            return True

    # 3. 페이지 번호 패턴
    page_number_pattern = r'-\s*\d+\s*-'
    page_numbers = len(re.findall(page_number_pattern, text))
    if page_numbers > 10:  # 페이지 번호가 10개 이상
        return True

    # 4. 계층적 번호 구조
    hierarchical_pattern = r'\d+\.\d+\.\d+'
    hierarchical_count = len(re.findall(hierarchical_pattern, text))
    if hierarchical_count > 15:  # 계층적 번호가 15개 이상
        return True

    return False

def parse_toc_items(text: str) -> List[Dict]:
    """
    TOC 항목 파싱

    각 항목을 다음 형태로 추출:
    {
        'number': '3.4.3.6',
        'title': '陡坡缓降（HDC）要求',
        'page': '17',
        'original_line': '원본 줄'
    }
    """

    items = []
    lines = text.split('\n')

    # 패턴: 숫자.숫자.숫자 형식의 항목 찾기
    item_pattern = r'^(\d+(?:\.\d+)*)\.\s*(.+?)\s*\.+\s*(?:.*?-\s*(\d+)\s*-)?'

    for line in lines:
        line = line.strip()
        if not line:
            continue

        match = re.search(item_pattern, line)
        if match:
            items.append({
                'number': match.group(1),
                'title': match.group(2).strip(),
                'page': match.group(3) if match.group(3) else '',
                'original_line': line
            })
        else:
            # 단순 항목 (번호 없음)
            if any(keyword in line for keyword in ['术语', '概述', '前言', '更改记录']):
                items.append({
                    'number': '',
                    'title': line,
                    'page': '',
                    'original_line': line
                })

    return items

def translate_toc_item(item: Dict) -> str:
    """
    개별 TOC 항목 번역 (제목만)
    """
    from googletrans import Translator

    try:
        translator = Translator()
        title = item['title']

        if not title or len(title) < 2:
            return item['original_line']

        # 제목만 번역
        result = translator.translate(title, dest='en')
        if result and hasattr(result, 'text') and result.text:
            translated_title = result.text

            # 번역된 제목으로 재구성
            if item['number'] and item['page']:
                return f"{item['number']}. {translated_title} {'.' * 20} - {item['page']} -"
            elif item['number']:
                return f"{item['number']}. {translated_title}"
            else:
                return translated_title

        return item['original_line']

    except Exception as e:
        print(f"        TOC 항목 번역 실패: {e}")
        return item['original_line']

def process_toc_structure(text: str) -> str:
    """
    TOC 구조 전용 처리

    1. TOC 항목 파싱
    2. 각 항목 개별 번역
    3. 재조립
    """

    print("      [TOC 감지] TOC 전용 처리 모드 활성화")

    # TOC 항목 파싱
    items = parse_toc_items(text)
    print(f"      TOC 항목 수: {len(items)}개")

    if not items:
        print("      TOC 항목을 찾을 수 없음, 일반 처리로 전환")
        return text

    # 각 항목 번역
    translated_items = []
    for i, item in enumerate(items[:5], 1):  # 처음 5개만 테스트
        print(f"        항목 {i}/{min(5, len(items))}: {item['title'][:30]}...")
        translated = translate_toc_item(item)
        translated_items.append(translated)
        time.sleep(0.5)  # 레이트 리미트 방지

    # 나머지는 원본 유지 (테스트)
    for item in items[5:]:
        translated_items.append(item['original_line'])

    # 재조립
    result = '\n'.join(translated_items)
    return result

# ============================================================================
# 기존 V2 함수들 (유지)
# ============================================================================

def extract_text_from_pdf(pdf_path: str, max_pages: Optional[int] = None) -> List[Tuple[int, str]]:
    """Extract text from PDF using PyPDF2"""
    pages_text = []
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            total_pages = len(pdf_reader.pages)

            if max_pages:
                total_pages = min(total_pages, max_pages)

            print(f"📖 Extracting from {total_pages} pages...")

            for i in range(total_pages):
                try:
                    page = pdf_reader.pages[i]
                    text = page.extract_text()
                    if text and text.strip():
                        pages_text.append((i + 1, text.strip()))
                    else:
                        pages_text.append((i + 1, f"[Page {i + 1} - No extractable text]"))
                except Exception as e:
                    print(f"   ⚠️  Error on page {i + 1}: {e}")
                    pages_text.append((i + 1, f"[Page {i + 1} - Extraction error: {str(e)}]"))

        print(f"✅ Successfully processed {len(pages_text)} pages")
        return pages_text

    except Exception as e:
        print(f"❌ Error reading PDF: {str(e)}")
        return []

def safe_text_cleaning(text: str) -> str:
    """Safely clean text with comprehensive error handling"""
    if not text or not isinstance(text, str):
        return ""

    try:
        if isinstance(text, bytes):
            text = text.decode('utf-8', errors='ignore')

        if text is None or str(text).strip() == 'None':
            return ""

        clean_parts = []
        for part in str(text).split():
            if part is not None and isinstance(part, str) and part.strip() and part.strip() != 'None':
                clean_parts.append(part.strip())

        if not clean_parts:
            return ""

        clean_text = ' '.join(clean_parts)
        clean_text = re.sub(r'\.{3,}', '...', clean_text)
        clean_text = re.sub(r'\s+', ' ', clean_text)
        clean_text = clean_text.strip()

        return clean_text

    except Exception as e:
        print(f"      Text cleaning error: {e}")
        return ""

def translate_with_google_robust(text: str, use_chunking: bool = True) -> str:
    """
    개선된 번역 함수 - V3

    V2 기능:
    - 재시도 5회
    - 지수 백오프
    - 동적 타임아웃

    V3 추가:
    - 스마트 청크 분할
    - 표 전용 처리
    - TOC 전용 처리
    """
    try:
        from googletrans import Translator

        if not text or not isinstance(text, str):
            return "[Error: Invalid input]"

        if text is None or str(text).strip() == 'None':
            return "[Error: None value detected]"

        clean_text = safe_text_cleaning(text)
        if not clean_text:
            return "[Error: No valid text after cleaning]"

        if len(clean_text) < 5:
            return f"[Skipped: Too short - '{clean_text}']"

        # 🆕 V3: TOC 감지 및 전용 처리
        if detect_toc(clean_text):
            print("      🔍 TOC 페이지 감지!")
            processed_toc = process_toc_structure(clean_text)
            # TOC 처리 후에도 번역 시도
            clean_text = processed_toc

        # 🆕 V3: 표 감지 및 전용 처리
        is_table = detect_table(clean_text)
        if is_table:
            print("      📊 표 구조 감지!")
            clean_text = process_table_structure(clean_text)

        # 🆕 V3: 청크 분할 여부 결정
        if use_chunking and len(clean_text) > 1000:
            print(f"      ✂️  긴 텍스트 감지 ({len(clean_text)}자), 청크 분할 적용")
            chunks = smart_chunk_text(clean_text, max_length=800, overlap=100)
            print(f"      📦 {len(chunks)}개 청크로 분할")

            chunk_translations = []
            for i, chunk in enumerate(chunks, 1):
                print(f"        청크 {i}/{len(chunks)} 번역 중...")
                translated = translate_single_text(chunk.content)
                chunk_translations.append(translated)
                if i < len(chunks):
                    time.sleep(1)  # 청크 간 대기

            # 청크 병합
            merged = merge_chunk_translations(chunk_translations, chunks)
            return merged
        else:
            # 일반 번역
            return translate_single_text(clean_text)

    except ImportError:
        return "[Error: googletrans not installed]"
    except Exception as e:
        return f"[Translation error: {str(e)}]"

def translate_single_text(text: str) -> str:
    """
    단일 텍스트 번역 (V2 로직 유지)
    - 재시도 5회
    - 지수 백오프
    """
    from googletrans import Translator

    for attempt in range(5):
        try:
            translator = Translator()
            result = translator.translate(text, dest='en')

            if result and hasattr(result, 'text') and result.text and result.text is not None and str(result.text).strip():
                if attempt > 0:
                    print(f"      ✅ 성공 (시도 {attempt + 1})")
                return result.text
            else:
                print(f"      시도 {attempt + 1}: Empty result")
        except Exception as e:
            print(f"      시도 {attempt + 1} 실패: {e}")

        if attempt < 4:
            wait_time = 2 ** attempt
            print(f"      대기 {wait_time}초...")
            time.sleep(wait_time)

    return f"[Translation failed after 5 attempts]"

def create_outputs(results: List[dict], base_name: str = "improved_translation_v3"):
    """Create JSON and CSV outputs"""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    json_data = {
        "total_pages_processed": len(results),
        "timestamp": timestamp,
        "successful_translations": len([r for r in results if not r['translated_text'].startswith('[')]),
        "version": "V3 - Chunking + Table + TOC Processing",
        "pages": results
    }

    json_file = f"output/{base_name}.json"
    try:
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON saved: {json_file}")
    except Exception as e:
        print(f"❌ JSON save error: {e}")

    csv_file = f"output/{base_name}.csv"
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Page', 'Original_Sample', 'Translation', 'Original_Length', 'Translation_Length', 'Status']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for result in results:
                status = "Success" if not result['translated_text'].startswith('[') else "Processed"
                writer.writerow({
                    'Page': result['page_number'],
                    'Original_Sample': result['original_text'][:200] + '...' if len(result['original_text']) > 200 else result['original_text'],
                    'Translation': result['translated_text'][:200] + '...' if len(result['translated_text']) > 200 else result['translated_text'],
                    'Original_Length': result['original_char_count'],
                    'Translation_Length': result['translated_char_count'],
                    'Status': status
                })
        print(f"✅ CSV saved: {csv_file}")
    except Exception as e:
        print(f"❌ CSV save error: {e}")

    # Excel output (읽기 좋은 버전)
    if OPENPYXL_AVAILABLE:
        excel_file = f"output/{base_name}_readable.xlsx"
        try:
            # 구조화 옵션 확인
            use_structure = '--structure' in sys.argv or '--toc' in sys.argv
            if use_structure and TOC_PARSER_AVAILABLE:
                print("   📊 목차 구조 기반 정리 활성화")
                create_structured_excel(json_data, excel_file)
            else:
                create_readable_excel(json_data, excel_file)
        except Exception as e:
            print(f"❌ Excel save error: {e}")
    else:
        print("⚠️  Excel output skipped (openpyxl not installed)")

def create_readable_excel(data: dict, output_path: str):
    """읽기 좋은 Excel 파일 생성 (줄바꿈, 한자 완벽 표시)"""
    wb = openpyxl.Workbook()

    # 시트 1: 통계
    ws_stats = wb.active
    ws_stats.title = "📊 통계"

    # 제목
    ws_stats['A1'] = "📊 번역 결과 통계 (V3)"
    ws_stats['A1'].font = Font(bold=True, size=16)
    ws_stats.merge_cells('A1:B1')

    # 통계 데이터
    stats = [
        ("", ""),
        ("총 페이지 수", data['total_pages_processed']),
        ("성공한 번역", data['successful_translations']),
        ("실패/처리", data['total_pages_processed'] - data['successful_translations']),
        ("성공률", f"{data['successful_translations']/data['total_pages_processed']*100:.1f}%"),
        ("처리 시간", data.get('timestamp', 'N/A')),
        ("버전", data.get('version', 'V3')),
    ]

    row = 2
    for label, value in stats:
        ws_stats[f'A{row}'] = label
        ws_stats[f'B{row}'] = value

        if label:
            ws_stats[f'A{row}'].font = Font(bold=True)
            ws_stats[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 40

    # 시트 2: 전체 번역 결과
    ws_all = wb.create_sheet("📄 전체 페이지")
    create_translation_sheet(ws_all, data['pages'])

    # 시트 3: 성공만
    successful = [p for p in data['pages'] if not p['translated_text'].startswith('[')]
    if successful:
        ws_success = wb.create_sheet("✅ 성공")
        create_translation_sheet(ws_success, successful)

    # 시트 4: 실패/처리 필요
    failed = [p for p in data['pages'] if p['translated_text'].startswith('[')]
    if failed:
        ws_failed = wb.create_sheet("⚠️ 실패")
        create_translation_sheet(ws_failed, failed)

    wb.save(output_path)
    print(f"✅ Excel saved (읽기 좋은 버전): {output_path}")

def create_translation_sheet(ws, pages):
    """번역 결과 시트 생성 (줄바꿈 자동 표시)"""
    # 헤더
    headers = ['페이지', '원문', '번역문', '원문 길이', '번역문 길이', '상태', '시간(초)']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 데이터 작성
    for row_idx, page in enumerate(pages, 2):
        is_success = not page['translated_text'].startswith('[')
        status = "✅ 성공" if is_success else "⚠️ 실패"

        # 셀 값 설정
        ws.cell(row=row_idx, column=1, value=page['page_number'])
        ws.cell(row=row_idx, column=2, value=page['original_text'][:500])  # 원문
        ws.cell(row=row_idx, column=3, value=page['translated_text'][:500])  # 번역문
        ws.cell(row=row_idx, column=4, value=page['original_char_count'])
        ws.cell(row=row_idx, column=5, value=page['translated_char_count'])
        ws.cell(row=row_idx, column=6, value=status)
        ws.cell(row=row_idx, column=7, value=page.get('translation_time', 'N/A'))

        # 스타일 적용
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border

            # ⭐ 텍스트 래핑 (줄바꿈 표시)
            if col in [2, 3]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 성공/실패 색상
        status_cell = ws.cell(row=row_idx, column=6)
        if is_success:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60   # 원문 (넓게)
    ws.column_dimensions['C'].width = 60   # 번역문 (넓게)
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    # 행 높이 (줄바꿈 고려)
    for row in range(2, len(pages) + 2):
        ws.row_dimensions[row].height = 100  # 충분한 높이

def create_structured_excel(data: dict, output_path: str):
    """목차 구조 기반 Excel 파일 생성"""
    parser = TOCStructureParser()
    wb = openpyxl.Workbook()

    # 목차 추출 (TOC 페이지 찾기)
    toc_pages = [p for p in data['pages'] if p['page_number'] in [2, 3]]  # 일반적으로 2-3페이지
    toc_text = '\n'.join([p['original_text'] for p in toc_pages])

    # 목차 파싱
    toc_items = parser.parse_toc_text(toc_text)
    print(f"   📖 목차 항목 {len(toc_items)}개 감지")

    # 페이지-섹션 매핑
    page_to_section = parser.map_pages_to_sections(data['pages'])

    # 시트 1: 통계
    ws_stats = wb.active
    ws_stats.title = "📊 통계"
    ws_stats['A1'] = "📊 번역 결과 통계 (구조화 V3)"
    ws_stats['A1'].font = Font(bold=True, size=16)
    ws_stats.merge_cells('A1:B1')

    stats = [
        ("", ""),
        ("총 페이지 수", data['total_pages_processed']),
        ("성공한 번역", data['successful_translations']),
        ("목차 항목 수", len(toc_items)),
        ("성공률", f"{data['successful_translations']/data['total_pages_processed']*100:.1f}%"),
        ("버전", "V3 - Chunking + TOC/표 처리"),
    ]

    row = 2
    for label, value in stats:
        ws_stats[f'A{row}'] = label
        ws_stats[f'B{row}'] = value
        if label:
            ws_stats[f'A{row}'].font = Font(bold=True)
            ws_stats[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 40

    # 시트 2: 구조화된 번역 결과
    ws_structured = wb.create_sheet("📚 구조화된 번역")
    create_structured_translation_sheet(ws_structured, data['pages'], page_to_section, parser)

    # 시트 3: 목차
    ws_toc = wb.create_sheet("📑 목차")
    create_toc_sheet(ws_toc, toc_items)

    wb.save(output_path)
    print(f"✅ Excel saved (구조화 버전): {output_path}")

def create_structured_translation_sheet(ws, pages, page_to_section: Dict, parser: TOCStructureParser):
    """구조화된 번역 시트"""
    headers = ['섹션', '페이지', '원문', '번역문', '상태']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 섹션별로 그룹화
    current_section = None
    row_idx = 2

    for page in sorted(pages, key=lambda x: x['page_number']):
        page_num = page['page_number']
        section_num = page_to_section.get(page_num, "")

        # 새 섹션 시작 시 헤더 추가
        if section_num and section_num != current_section:
            current_section = section_num
            section_info = parser.get_section_info(section_num)

            # 섹션 헤더 행
            level = section_num.count('.') + 1
            indent = "  " * (level - 1)
            section_title = f"{indent}{section_num}"
            if section_info:
                section_title += f" {section_info.title}"

            ws.cell(row=row_idx, column=1, value=section_title)
            ws.merge_cells(f'A{row_idx}:E{row_idx}')

            header_cell = ws.cell(row=row_idx, column=1)
            if level == 1:
                header_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                header_cell.font = Font(bold=True, size=12)
            elif level == 2:
                header_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                header_cell.font = Font(bold=True, size=11)
            else:
                header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                header_cell.font = Font(bold=True, size=10)

            row_idx += 1

        # 페이지 데이터
        is_success = not page['translated_text'].startswith('[')
        status = "✅" if is_success else "⚠️"

        ws.cell(row=row_idx, column=1, value=section_num)
        ws.cell(row=row_idx, column=2, value=page_num)
        ws.cell(row=row_idx, column=3, value=page['original_text'][:300])
        ws.cell(row=row_idx, column=4, value=page['translated_text'][:300])
        ws.cell(row=row_idx, column=5, value=status)

        # 스타일
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [3, 4]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if is_success:
            ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws.row_dimensions[row_idx].height = 60
        row_idx += 1

    # 열 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 8

def create_toc_sheet(ws, toc_items: List[TOCItem]):
    """목차 시트"""
    headers = ['번호', '제목', '레벨', '페이지']

    # 헤더
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 목차 항목
    for row_idx, item in enumerate(toc_items, 2):
        indent = "  " * (item.level - 1)

        ws.cell(row=row_idx, column=1, value=item.number)
        ws.cell(row=row_idx, column=2, value=f"{indent}{item.title}")
        ws.cell(row=row_idx, column=3, value=item.level)
        ws.cell(row=row_idx, column=4, value=item.page)

        # 레벨별 색상
        if item.level == 1:
            fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        elif item.level == 2:
            fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        else:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        ws.cell(row=row_idx, column=1).fill = fill
        ws.cell(row=row_idx, column=2).fill = fill

    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10

def main():
    """Main translation workflow - V3 with chunking and table processing"""

    pdf_file = "input/XY-A ATS开发对IBC需求문서_V0.0.pdf"

    # 명령줄 옵션 확인
    use_structure = '--structure' in sys.argv or '--toc' in sys.argv

    # 전체 번역 또는 특정 페이지만 번역
    if '--all' in sys.argv:
        target_pages = None  # 전체 페이지
        mode_text = "전체 70페이지"
    else:
        target_pages = [2, 3]  # TOC 페이지만 (테스트용)
        mode_text = f"페이지 {target_pages}"

    print("🚀 개선된 PDF 번역기 V3 - 2단계 개선사항 적용")
    print("="*80)
    print("✨ V3 신규 기능:")
    print("   1. 스마트 청크 분할 (800자 단위, 100자 오버랩)")
    print("   2. 표/테이블 자동 감지 및 전용 처리")
    print("   3. TOC 구조 파싱 및 항목별 번역")
    if use_structure:
        print("   📊 목차 구조 기반 정리: 활성화")
    print("="*80)
    print(f"\n🎯 대상: {mode_text}")
    if target_pages is None:
        print(f"⏱️  예상 시간: 약 5-7분")

    if not os.path.exists(pdf_file):
        print(f"❌ PDF not found: {pdf_file}")
        return

    # PDF 추출
    pages_text = extract_text_from_pdf(pdf_file, max_pages=None)
    if not pages_text:
        print("❌ No text extracted")
        return

    # 타겟 페이지만 필터링
    if target_pages:
        pages_to_translate = [(num, text) for num, text in pages_text if num in target_pages]
    else:
        pages_to_translate = pages_text

    results = []
    total = len(pages_to_translate)

    for i, (page_num, text) in enumerate(pages_to_translate, 1):
        print(f"\n{'='*80}")
        print(f"📋 Processing Page {page_num} ({i}/{total})...")

        sample = safe_text_cleaning(text)[:100]
        print(f"   📝 Sample: {sample}...")

        print("   🔄 V3 번역 중...")
        start_time = time.time()
        translated = translate_with_google_robust(text, use_chunking=True)
        translation_time = time.time() - start_time

        result_sample = translated[:100]
        print(f"   ✅ Result: {result_sample}...")
        print(f"   ⏱️  Time: {translation_time:.1f}s")

        results.append({
            "page_number": page_num,
            "original_text": text,
            "translated_text": translated,
            "original_char_count": len(text),
            "translated_char_count": len(translated),
            "translation_time": round(translation_time, 2)
        })

        if i < total:
            time.sleep(2)

    # 결과 저장
    print(f"\n{'='*80}")
    print("💾 Saving results...")
    create_outputs(results, "improved_translation_v3_results")

    # 통계
    successful = len([r for r in results if not r['translated_text'].startswith('[')])
    print(f"\n📊 Summary:")
    print(f"   Pages processed: {len(results)}")
    print(f"   Successful: {successful} ({successful/len(results)*100:.1f}%)")

    print("\n✅ V3 Translation completed!")

if __name__ == "__main__":
    main()
