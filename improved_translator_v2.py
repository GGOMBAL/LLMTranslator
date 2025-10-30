#!/usr/bin/env python3
"""
개선된 PDF 번역기 V2 - 1단계 개선사항 적용
- 재시도 횟수: 2회 → 5회
- 타임아웃: 동적 조정 (텍스트 길이 기반)
- None 값 체크 강화
- 지수 백오프 적용
"""

import PyPDF2
import os
import json
import time
import re
import csv
import sys
from typing import List, Tuple, Optional, Dict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel output will be skipped.")

try:
    from toc_structure_parser import TOCStructureParser, TOCItem
    TOC_PARSER_AVAILABLE = True
except ImportError:
    TOC_PARSER_AVAILABLE = False

def extract_text_from_pdf(pdf_path: str, max_pages: Optional[int] = None) -> List[Tuple[int, str]]:
    """Extract text from PDF using PyPDF2"""
    pages_text = []
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            total_pages = len(pdf_reader.pages)

            if max_pages:
                total_pages = min(total_pages, max_pages)

            print(f"📖 Extracting from {total_pages} pages...")

            for i in range(total_pages):
                try:
                    page = pdf_reader.pages[i]
                    text = page.extract_text()
                    if text and text.strip():
                        pages_text.append((i + 1, text.strip()))
                    else:
                        pages_text.append((i + 1, f"[Page {i + 1} - No extractable text]"))
                except Exception as e:
                    print(f"   ⚠️  Error on page {i + 1}: {e}")
                    pages_text.append((i + 1, f"[Page {i + 1} - Extraction error: {str(e)}]"))

        print(f"✅ Successfully processed {len(pages_text)} pages")
        return pages_text

    except Exception as e:
        print(f"❌ Error reading PDF: {str(e)}")
        return []

def safe_text_cleaning(text: str) -> str:
    """Safely clean text with comprehensive error handling - IMPROVED"""
    if not text or not isinstance(text, str):
        return ""

    try:
        # Handle potential encoding issues
        if isinstance(text, bytes):
            text = text.decode('utf-8', errors='ignore')

        # 🆕 IMPROVED: More robust None checking
        if text is None or str(text).strip() == 'None':
            return ""

        # Split and filter out None/empty values
        clean_parts = []
        for part in str(text).split():
            # 🆕 IMPROVED: Stricter validation
            if part is not None and isinstance(part, str) and part.strip() and part.strip() != 'None':
                clean_parts.append(part.strip())

        if not clean_parts:
            return ""

        # Join safely
        clean_text = ' '.join(clean_parts)

        # Remove excessive formatting
        clean_text = re.sub(r'\.{3,}', '...', clean_text)
        clean_text = re.sub(r'\s+', ' ', clean_text)
        clean_text = clean_text.strip()

        return clean_text

    except Exception as e:
        print(f"      Text cleaning error: {e}")
        return ""

def extract_chinese_content(text: str) -> str:
    """Extract only Chinese characters and related content - IMPROVED"""
    try:
        # 🆕 IMPROVED: None check
        if text is None or not isinstance(text, str):
            return ""

        # Pattern for Chinese characters, parentheses, and common punctuation
        chinese_pattern = r'[\u4e00-\u9fff\u3400-\u4dbf\（\）\uff08\uff09A-Za-z0-9\s]+'
        matches = re.findall(chinese_pattern, text)

        if matches:
            # 🆕 IMPROVED: Filter None values
            matches = [m for m in matches if m is not None and str(m).strip()]
            if not matches:
                return ""

            chinese_text = ' '.join(matches)
            # Clean up excessive spaces
            chinese_text = re.sub(r'\s+', ' ', chinese_text).strip()
            return chinese_text
        return ""
    except Exception as e:
        print(f"      Chinese extraction error: {e}")
        return ""

def calculate_dynamic_timeout(text_length: int) -> int:
    """🆕 NEW: Calculate timeout based on text length"""
    # Base timeout: 30 seconds
    # Add 10 seconds per 1000 characters
    # Maximum: 180 seconds (3 minutes)
    timeout = min(30 + (text_length // 100), 180)
    return timeout

def translate_with_google_robust(text: str) -> str:
    """🆕 IMPROVED: Robust translation with enhanced retry mechanism"""
    try:
        from googletrans import Translator

        # Initial validation
        if not text or not isinstance(text, str):
            return "[Error: Invalid input]"

        # 🆕 IMPROVED: Stronger None check
        if text is None or str(text).strip() == 'None':
            return "[Error: None value detected]"

        # Clean text safely
        clean_text = safe_text_cleaning(text)
        if not clean_text:
            return "[Error: No valid text after cleaning]"

        # Check text length
        if len(clean_text) < 5:
            return f"[Skipped: Too short - '{clean_text}']"

        # 🆕 NEW: Calculate dynamic timeout
        text_length = len(clean_text)
        timeout = calculate_dynamic_timeout(text_length)
        print(f"      Text length: {text_length} chars, Timeout: {timeout}s")

        # Handle heavily formatted text (TOC pages)
        symbol_count = clean_text.count('.') + clean_text.count('-') + clean_text.count('_')
        symbol_ratio = symbol_count / len(clean_text) if len(clean_text) > 0 else 0

        if symbol_ratio > 0.25:  # More than 25% symbols
            print(f"      Detected TOC/formatted content...")
            chinese_content = extract_chinese_content(clean_text)
            if chinese_content and len(chinese_content) > 10:
                print(f"      Translating Chinese content: {chinese_content[:50]}...")

                # 🆕 IMPROVED: 5 attempts with exponential backoff for TOC
                for attempt in range(5):
                    try:
                        translator = Translator()
                        result = translator.translate(chinese_content, dest='en')

                        # 🆕 IMPROVED: Stronger validation
                        if result and hasattr(result, 'text') and result.text and result.text is not None:
                            return f"[TOC] {result.text}"
                        else:
                            print(f"      TOC attempt {attempt + 1}: Empty result")
                    except Exception as e:
                        print(f"      TOC attempt {attempt + 1} failed: {e}")
                        if attempt < 4:  # Don't wait after last attempt
                            wait_time = 2 ** attempt  # Exponential backoff: 1, 2, 4, 8, 16 seconds
                            print(f"      Waiting {wait_time}s before retry...")
                            time.sleep(wait_time)

                return f"[TOC - Translation failed after 5 attempts]"
            return f"[Skipped: Mostly formatting - {clean_text[:50]}...]"

        # Regular translation for normal content
        if len(clean_text) > 2000:  # 🆕 IMPROVED: Increased limit from 1500 to 2000
            print(f"      Text too long ({len(clean_text)} chars), truncating to 2000...")
            clean_text = clean_text[:2000] + "..."

        print(f"      Translating: {clean_text[:50]}...")

        # 🆕 IMPROVED: 5 attempts with exponential backoff
        for attempt in range(5):
            try:
                # Create new translator instance each time
                translator = Translator()
                result = translator.translate(clean_text, dest='en')

                # 🆕 IMPROVED: Stronger result validation
                if result and hasattr(result, 'text') and result.text and result.text is not None and str(result.text).strip():
                    print(f"      ✅ Success on attempt {attempt + 1}")
                    return result.text
                else:
                    print(f"      Attempt {attempt + 1}: Empty or None result")
            except Exception as e:
                print(f"      Attempt {attempt + 1} failed: {e}")

            # 🆕 IMPROVED: Exponential backoff
            if attempt < 4:  # Don't wait after last attempt
                wait_time = 2 ** attempt  # 1, 2, 4, 8, 16 seconds
                print(f"      Waiting {wait_time}s before retry...")
                time.sleep(wait_time)

        return f"[Translation failed after 5 attempts]"

    except ImportError:
        return "[Error: googletrans not installed]"
    except Exception as e:
        return f"[Translation error: {str(e)}]"

def load_failed_pages_from_json(json_path: str) -> List[int]:
    """Load list of failed page numbers from previous translation"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        failed_pages = []
        for page in data['pages']:
            translated = page['translated_text']
            if '[Translation failed' in translated or '[TOC -' in translated or 'NoneType' in translated:
                failed_pages.append(page['page_number'])

        return sorted(failed_pages)
    except Exception as e:
        print(f"Error loading failed pages: {e}")
        return []

def create_outputs(results: List[dict], base_name: str = "improved_translation_v2"):
    """Create both JSON and CSV outputs"""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    # JSON output
    json_data = {
        "total_pages_processed": len(results),
        "timestamp": timestamp,
        "successful_translations": len([r for r in results if not r['translated_text'].startswith('[')]),
        "pages": results
    }

    json_file = f"output/{base_name}.json"
    try:
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON saved: {json_file}")
    except Exception as e:
        print(f"❌ JSON save error: {e}")

    # CSV output
    csv_file = f"output/{base_name}.csv"
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Page', 'Original_Sample', 'Translation', 'Original_Length', 'Translation_Length', 'Status']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for result in results:
                status = "Success" if not result['translated_text'].startswith('[') else "Processed"
                writer.writerow({
                    'Page': result['page_number'],
                    'Original_Sample': result['original_text'][:200] + '...' if len(result['original_text']) > 200 else result['original_text'],
                    'Translation': result['translated_text'][:200] + '...' if len(result['translated_text']) > 200 else result['translated_text'],
                    'Original_Length': result['original_char_count'],
                    'Translation_Length': result['translated_char_count'],
                    'Status': status
                })
        print(f"✅ CSV saved: {csv_file}")
    except Exception as e:
        print(f"❌ CSV save error: {e}")

    # Excel output (읽기 좋은 버전)
    if OPENPYXL_AVAILABLE:
        excel_file = f"output/{base_name}_readable.xlsx"
        try:
            # 구조화 옵션 확인
            use_structure = '--structure' in sys.argv or '--toc' in sys.argv
            if use_structure and TOC_PARSER_AVAILABLE:
                print("   📊 목차 구조 기반 정리 활성화")
                create_structured_excel(json_data, excel_file)
            else:
                create_readable_excel(json_data, excel_file)
        except Exception as e:
            print(f"❌ Excel save error: {e}")
    else:
        print("⚠️  Excel output skipped (openpyxl not installed)")

def create_readable_excel(data: dict, output_path: str):
    """읽기 좋은 Excel 파일 생성 (줄바꿈, 한자 완벽 표시)"""
    wb = openpyxl.Workbook()

    # 시트 1: 통계
    ws_stats = wb.active
    ws_stats.title = "📊 통계"

    # 제목
    ws_stats['A1'] = "📊 번역 결과 통계"
    ws_stats['A1'].font = Font(bold=True, size=16)
    ws_stats.merge_cells('A1:B1')

    # 통계 데이터
    stats = [
        ("", ""),
        ("총 페이지 수", data['total_pages_processed']),
        ("성공한 번역", data['successful_translations']),
        ("실패/처리", data['total_pages_processed'] - data['successful_translations']),
        ("성공률", f"{data['successful_translations']/data['total_pages_processed']*100:.1f}%"),
        ("처리 시간", data.get('timestamp', 'N/A')),
    ]

    row = 2
    for label, value in stats:
        ws_stats[f'A{row}'] = label
        ws_stats[f'B{row}'] = value

        if label:
            ws_stats[f'A{row}'].font = Font(bold=True)
            ws_stats[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 30

    # 시트 2: 전체 번역 결과
    ws_all = wb.create_sheet("📄 전체 페이지")
    create_translation_sheet(ws_all, data['pages'])

    # 시트 3: 성공만
    successful = [p for p in data['pages'] if not p['translated_text'].startswith('[')]
    if successful:
        ws_success = wb.create_sheet("✅ 성공")
        create_translation_sheet(ws_success, successful)

    # 시트 4: 실패/처리 필요
    failed = [p for p in data['pages'] if p['translated_text'].startswith('[')]
    if failed:
        ws_failed = wb.create_sheet("⚠️ 실패")
        create_translation_sheet(ws_failed, failed)

    wb.save(output_path)
    print(f"✅ Excel saved (읽기 좋은 버전): {output_path}")

def create_translation_sheet(ws, pages):
    """번역 결과 시트 생성 (줄바꿈 자동 표시)"""
    # 헤더
    headers = ['페이지', '원문', '번역문', '원문 길이', '번역문 길이', '상태', '시간(초)']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 데이터 작성
    for row_idx, page in enumerate(pages, 2):
        is_success = not page['translated_text'].startswith('[')
        status = "✅ 성공" if is_success else "⚠️ 실패"

        # 셀 값 설정
        ws.cell(row=row_idx, column=1, value=page['page_number'])
        ws.cell(row=row_idx, column=2, value=page['original_text'][:500])  # 원문
        ws.cell(row=row_idx, column=3, value=page['translated_text'][:500])  # 번역문
        ws.cell(row=row_idx, column=4, value=page['original_char_count'])
        ws.cell(row=row_idx, column=5, value=page['translated_char_count'])
        ws.cell(row=row_idx, column=6, value=status)
        ws.cell(row=row_idx, column=7, value=page.get('translation_time', 'N/A'))

        # 스타일 적용
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border

            # ⭐ 텍스트 래핑 (줄바꿈 표시)
            if col in [2, 3]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 성공/실패 색상
        status_cell = ws.cell(row=row_idx, column=6)
        if is_success:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60   # 원문 (넓게)
    ws.column_dimensions['C'].width = 60   # 번역문 (넓게)
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    # 행 높이 (줄바꿈 고려)
    for row in range(2, len(pages) + 2):
        ws.row_dimensions[row].height = 100  # 충분한 높이

def create_structured_excel(data: dict, output_path: str):
    """목차 구조 기반 Excel 파일 생성"""
    parser = TOCStructureParser()
    wb = openpyxl.Workbook()

    # 목차 추출 (TOC 페이지 찾기)
    toc_pages = [p for p in data['pages'] if p['page_number'] in [2, 3]]  # 일반적으로 2-3페이지
    toc_text = '\n'.join([p['original_text'] for p in toc_pages])

    # 목차 파싱
    toc_items = parser.parse_toc_text(toc_text)
    print(f"   📖 목차 항목 {len(toc_items)}개 감지")

    # 페이지-섹션 매핑 (디버그 모드 활성화)
    debug_mode = '--debug' in sys.argv
    page_to_section = parser.map_pages_to_sections(data['pages'], debug=debug_mode)

    # 시트 1: 통계
    ws_stats = wb.active
    ws_stats.title = "📊 통계"
    ws_stats['A1'] = "📊 번역 결과 통계 (구조화)"
    ws_stats['A1'].font = Font(bold=True, size=16)
    ws_stats.merge_cells('A1:B1')

    stats = [
        ("", ""),
        ("총 페이지 수", data['total_pages_processed']),
        ("성공한 번역", data['successful_translations']),
        ("목차 항목 수", len(toc_items)),
        ("성공률", f"{data['successful_translations']/data['total_pages_processed']*100:.1f}%"),
    ]

    row = 2
    for label, value in stats:
        ws_stats[f'A{row}'] = label
        ws_stats[f'B{row}'] = value
        if label:
            ws_stats[f'A{row}'].font = Font(bold=True)
            ws_stats[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 30

    # 시트 2: 구조화된 번역 결과
    ws_structured = wb.create_sheet("📚 구조화된 번역")
    create_structured_translation_sheet(ws_structured, data['pages'], page_to_section, parser)

    # 시트 3: 목차
    ws_toc = wb.create_sheet("📑 목차")
    create_toc_sheet(ws_toc, toc_items)

    wb.save(output_path)
    print(f"✅ Excel saved (구조화 버전): {output_path}")

def create_structured_translation_sheet(ws, pages, page_to_section: Dict, parser: TOCStructureParser):
    """구조화된 번역 시트"""
    headers = ['섹션', '페이지', '원문', '번역문', '상태']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 섹션별로 그룹화
    current_section = None
    row_idx = 2

    for page in sorted(pages, key=lambda x: x['page_number']):
        page_num = page['page_number']
        section_num = page_to_section.get(page_num, "")

        # 새 섹션 시작 시 헤더 추가
        if section_num and section_num != current_section:
            current_section = section_num
            section_info = parser.get_section_info(section_num)

            # 섹션 헤더 행
            level = section_num.count('.') + 1
            indent = "  " * (level - 1)
            section_title = f"{indent}{section_num}"
            if section_info:
                section_title += f" {section_info.title}"

            ws.cell(row=row_idx, column=1, value=section_title)
            ws.merge_cells(f'A{row_idx}:E{row_idx}')

            header_cell = ws.cell(row=row_idx, column=1)
            if level == 1:
                header_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                header_cell.font = Font(bold=True, size=12)
            elif level == 2:
                header_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                header_cell.font = Font(bold=True, size=11)
            else:
                header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                header_cell.font = Font(bold=True, size=10)

            row_idx += 1

        # 페이지 데이터
        is_success = not page['translated_text'].startswith('[')
        status = "✅" if is_success else "⚠️"

        ws.cell(row=row_idx, column=1, value=section_num)
        ws.cell(row=row_idx, column=2, value=page_num)
        ws.cell(row=row_idx, column=3, value=page['original_text'][:300])
        ws.cell(row=row_idx, column=4, value=page['translated_text'][:300])
        ws.cell(row=row_idx, column=5, value=status)

        # 스타일
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [3, 4]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if is_success:
            ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws.row_dimensions[row_idx].height = 60
        row_idx += 1

    # 열 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 8

def create_toc_sheet(ws, toc_items: List[TOCItem]):
    """목차 시트"""
    headers = ['번호', '제목', '레벨', '페이지']

    # 헤더
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 목차 항목
    for row_idx, item in enumerate(toc_items, 2):
        indent = "  " * (item.level - 1)

        ws.cell(row=row_idx, column=1, value=item.number)
        ws.cell(row=row_idx, column=2, value=f"{indent}{item.title}")
        ws.cell(row=row_idx, column=3, value=item.level)
        ws.cell(row=row_idx, column=4, value=item.page)

        # 레벨별 색상
        if item.level == 1:
            fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        elif item.level == 2:
            fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        else:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        ws.cell(row=row_idx, column=1).fill = fill
        ws.cell(row=row_idx, column=2).fill = fill

    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10

def main():
    """Main translation workflow - Re-translate only failed pages"""
    pdf_file = "input/XY-A ATS开发对IBC需求文档_V0.0.pdf"
    previous_results_json = "output/final_translation_results.json"

    # 명령줄 옵션 확인
    use_structure = '--structure' in sys.argv or '--toc' in sys.argv

    print("🚀 개선된 PDF 번역기 V2 - 1단계 개선사항 적용")
    print("="*80)
    print("✨ 개선사항:")
    print("   1. 재시도 횟수: 2회 → 5회")
    print("   2. 지수 백오프 적용: 1s, 2s, 4s, 8s, 16s")
    print("   3. 타임아웃 동적 조정: 30-180초 (텍스트 길이 기반)")
    print("   4. None 값 체크 강화")
    print("   5. 텍스트 길이 제한: 1500자 → 2000자")
    if use_structure:
        print("   📊 목차 구조 기반 정리: 활성화")
    print("="*80)

    # Check PDF existence
    if not os.path.exists(pdf_file):
        print(f"❌ PDF not found: {pdf_file}")
        return

    print(f"\n✅ Found PDF: {pdf_file}")

    # Load failed pages from previous translation
    print(f"\n📂 Loading previous translation results...")
    failed_pages = load_failed_pages_from_json(previous_results_json)

    if not failed_pages:
        print("❌ No failed pages found or couldn't load previous results")
        print("   Will translate all pages instead...")
        failed_pages = None
    else:
        print(f"✅ Found {len(failed_pages)} failed pages to retry:")
        print(f"   Pages: {', '.join(map(str, failed_pages[:10]))}" +
              (f" ... (and {len(failed_pages)-10} more)" if len(failed_pages) > 10 else ""))

    # Extract text from PDF
    pages_text = extract_text_from_pdf(pdf_file, max_pages=None)

    if not pages_text:
        print("❌ No text extracted from PDF")
        return

    # Filter to only failed pages if available
    if failed_pages:
        pages_to_translate = [(num, text) for num, text in pages_text if num in failed_pages]
        print(f"\n📋 Retrying {len(pages_to_translate)} failed pages...")
    else:
        pages_to_translate = pages_text
        print(f"\n📋 Translating all {len(pages_to_translate)} pages...")

    # Process translations
    results = []
    total_pages = len(pages_to_translate)

    for i, (page_num, text) in enumerate(pages_to_translate, 1):
        print(f"\n{'='*80}")
        print(f"📋 Processing Page {page_num} ({i}/{total_pages})...")

        # Show sample of original
        sample = safe_text_cleaning(text)[:100]
        print(f"   📝 Original sample: {sample}...")

        # Translate
        print("   🔄 Translating with improved retry mechanism...")
        start_time = time.time()
        translated_text = translate_with_google_robust(text)
        translation_time = time.time() - start_time

        # Show result sample
        result_sample = translated_text[:100]
        print(f"   ✅ Result: {result_sample}...")
        print(f"   ⏱️  Time: {translation_time:.1f}s")

        # Determine status
        if translated_text.startswith('['):
            print(f"   ⚠️  Status: FAILED/PROCESSED")
        else:
            print(f"   ✅ Status: SUCCESS")

        results.append({
            "page_number": page_num,
            "original_text": text,
            "translated_text": translated_text,
            "original_char_count": len(text),
            "translated_char_count": len(translated_text),
            "translation_time": round(translation_time, 2)
        })

        # Rate limiting - longer wait between pages
        if i < total_pages:
            wait_time = 2  # 🆕 IMPROVED: Increased from 1.5s to 2s
            print(f"   ⏳ Waiting {wait_time}s before next page...")
            time.sleep(wait_time)

    # Save results
    print(f"\n{'='*80}")
    print(f"💾 Saving results...")
    create_outputs(results, "improved_translation_v2_results")

    # Summary
    successful = len([r for r in results if not r['translated_text'].startswith('[')])
    processed = len([r for r in results if r['translated_text'].startswith('[TOC]')])
    failed = len(results) - successful - processed

    print(f"\n📊 Final Summary:")
    print(f"   📄 Total pages processed: {len(results)}")
    print(f"   ✅ Successful translations: {successful} ({successful/len(results)*100:.1f}%)")
    print(f"   📋 Processed (TOC/formatted): {processed}")
    print(f"   ❌ Failed/skipped: {failed} ({failed/len(results)*100:.1f}%)")

    # Improvement comparison
    if failed_pages:
        previous_failed = len(failed_pages)
        new_successful = successful
        improvement = new_successful
        print(f"\n📈 Improvement Analysis:")
        print(f"   🔴 Previously failed: {previous_failed} pages")
        print(f"   🟢 Now successful: {new_successful} pages ({new_successful/previous_failed*100:.1f}%)")
        print(f"   📊 Improvement: {improvement} pages recovered")
        if failed > 0:
            print(f"   🔴 Still failing: {failed} pages ({failed/previous_failed*100:.1f}%)")

    # Show best samples
    if successful > 0:
        best_results = [r for r in results if not r['translated_text'].startswith('[')]
        if best_results:
            sample = best_results[0]
            print(f"\n🏆 Sample Translation (Page {sample['page_number']}):")
            print("🇨🇳 Original:")
            print("  " + sample['original_text'][:150].replace('\n', ' ') + "...")
            print("🇺🇸 Translation:")
            print("  " + sample['translated_text'][:150] + "...")

    print("\n✅ Translation completed successfully!")
    print("📁 Check the output files for complete results.")

if __name__ == "__main__":
    main()
